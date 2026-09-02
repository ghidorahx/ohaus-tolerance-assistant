#!/usr/bin/env python3
"""Build a deterministic, workbook-grounded retrieval evaluation fixture."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl


APP_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = APP_ROOT / "MMMDF_EN_US_20260605_AI_Organized 2.xlsx"
DEFAULT_OUTPUT = APP_ROOT / "data" / "master-retrieval-eval.json"
DEFAULT_PROFILE_OUTPUT = APP_ROOT / "data" / "master-retrieval-eval-profile.json"
SOURCE_SHEET = "Product_Catalog_AI"
PROFILE_VERSION = "2.0.0"
FINGERPRINT_SCHEMA_VERSION = "1.0.0"
RETRIEVAL_CODE_FILES = (
    "lib/master-catalog-rag.mjs",
    "app/api/sales/route.ts",
    "scripts/evaluate-master-retrieval.mjs",
    "scripts/verify-master-retrieval-profile.mjs",
    "scripts/build-master-retrieval-eval.py",
)
WRANGLER_DEPLOY_CONFIG = APP_ROOT / "wrangler.deploy.jsonc"

IDENTITY_FIELDS = [
    "Material Number",
    "Material Description (Global English)",
    "Parent Family Name",
    "Family Name",
    "Trade Name",
]
SIX_DIGIT_MATERIALS = ["214642", "923345", "923389"]
PLACEHOLDERS = {"n/a", "na", "not applicable", "none", "null", "tbd", "unknown", "-", "--"}

TECHNICAL_CASES = [
    ("Maximum Capacity {metric}", "maximum_capacity", "How much can {reference} weigh at most?"),
    ("Readability {metric}", "readability", "What is the smallest displayed increment for {reference}?"),
    ("Stabilization Time", "stabilization_time", "How quickly does {reference} stabilize?"),
    ("Battery Life", "battery_life", "How long is the listed battery life for {reference}?"),
    ("Power", "power", "What power source does {reference} use?"),
    ("Legal for Trade", "legal_for_trade", "Is {reference} listed as legal for trade?"),
    ("IP Rating", "ip_rating", "What ingress-protection rating is specified for {reference}?"),
    ("Maximum Speed", "maximum_speed", "What is the maximum operating speed of {reference}?"),
    ("Speed Range", "speed_range", "What speed range is specified for {reference}?"),
    ("Temperature Range {metric}", "temperature_range", "What metric temperature range is listed for {reference}?"),
    ("Measurement Range", "measurement_range", "What measurement range does {reference} cover?"),
    ("pH measuring range", "ph_range", "Across what pH range can {reference} measure?"),
    ("Conductivity measuring range", "conductivity_range", "What conductivity range is listed for {reference}?"),
    ("Accuracy", "accuracy", "What accuracy is specified for {reference}?"),
    ("Calibration Certificate", "calibration_certificate", "What calibration certificate is listed for {reference}?"),
    ("Weight Tolerance", "weight_tolerance", "What weight tolerance is specified for {reference}?"),
    ("NTEP Approval", "ntep_approval", "What does the catalog say about NTEP approval for {reference}?"),
    ("Pan Size {Width} {metric}", "pan_width_metric", "What is the metric pan width for {reference}?"),
    ("Platform Size {Length} {metric}", "platform_length_metric", "What is the metric platform length for {reference}?"),
    ("Working Environment {metric}", "working_environment", "What metric working environment is specified for {reference}?"),
]

RELATIONSHIP_CASES = [
    ("Relationship / Accessories", "accessories", "Which accessory material numbers are listed for {reference}?", "material_ids"),
    ("Relationship / Cross Selling", "cross_selling", "Which cross-sell material numbers are listed for {reference}?", "material_ids"),
    ("Relationship / Services", "services", "Which service codes are linked to {reference}?", "material_ids"),
    ("Relationship / Spare Parts", "spare_parts", "Which spare-part material numbers are listed for {reference}?", "material_ids"),
    ("Relationship / Upsellings", "upsellings", "Which upsell material numbers are listed for {reference}?", "material_ids"),
    ("Relationship / Replacements", "replacements", "Which replacement material numbers are listed for {reference}?", "material_ids"),
    ("Compatible Models", "compatible_models", "Which models are listed as compatible with {reference}?", "values"),
    ("Rotor Compatibility", "rotor_compatibility", "Which rotors are listed as compatible with {reference}?", "values"),
    ("Accessories Included", "included_accessories", "Which accessories are included with {reference}?", "values"),
    ("Sample Vials Compatibility", "sample_vial_compatibility", "Which sample-vial format is compatible with {reference}?", "values"),
]

DOCUMENT_FIELDS = [
    "EN Data Sheets 1",
    "EN Data Sheets 2",
    "EN Data Sheets 3",
    "EN User Guide 1",
    "EN User Guide 2",
    "EN Manuals 1",
    "EN Manuals 2",
]

UNSUPPORTED_CASES = [
    ("live_price", "What is today's price for material {material_number}?"),
    ("inventory_quantity", "How many units of material {material_number} are in stock right now?"),
    ("live_lead_time", "What is the current ship date and lead time for material {material_number}?"),
    ("current_discount", "What live discount can we offer on material {material_number} today?"),
    ("order_status", "Has the latest customer order for material {material_number} shipped yet?"),
    ("future_restock", "When will material {material_number} be back in inventory?"),
]

SEMANTIC_DISCOVERY_CASES = [
    (
        "cell-lysis-soil-biosafety-cabinet",
        "83041308",
        "What product would help me break open cells and homogenize soil samples inside a cramped biosafety cabinet?",
        ["Application", "Benefit Text 1"],
    ),
    (
        "quiet-safe-mini-centrifuge",
        "30134157",
        "I need a quiet mini centrifuge with imbalance protection that stops the rotor if the lid opens. What fits?",
        ["Benefit Text 1", "Benefit Text 2"],
    ),
    (
        "automatic-moisture-method",
        "30241165",
        "Which moisture tester can analyze a sample and automatically create the drying method for me?",
        ["Application", "Benefit Text 1"],
    ),
]


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\u00a0", " ").strip()


def normalized(value: object) -> str:
    return re.sub(r"\s+", " ", clean(value)).casefold()


def is_grounded_value(value: object) -> bool:
    value_text = clean(value)
    return bool(value_text) and normalized(value_text) not in PLACEHOLDERS


def slug(value: str) -> str:
    ascii_text = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "-", ascii_text.casefold()).strip("-")


def material_sort_key(material_number: str) -> tuple[int, str]:
    return len(material_number), material_number


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def canonical_sha256(value: object) -> str:
    payload = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def javascript_export(source: str, name: str) -> object:
    match = re.search(rf"export const {re.escape(name)}\s*=\s*([^;]+);", source)
    if not match:
        raise SystemExit(f"Cannot fingerprint retrieval build; missing JavaScript export {name}.")
    raw = match.group(1).strip()
    try:
        return json.loads(raw)
    except json.JSONDecodeError as error:
        raise SystemExit(f"Cannot fingerprint retrieval build; unsupported value for {name}: {raw}") from error


def retrieval_build_profile() -> dict[str, object]:
    code_files = [
        {"path": relative_path, "sha256": file_sha256(APP_ROOT / relative_path)}
        for relative_path in RETRIEVAL_CODE_FILES
    ]
    rag_source = (APP_ROOT / "lib/master-catalog-rag.mjs").read_text(encoding="utf-8")
    deploy_config = json.loads(WRANGLER_DEPLOY_CONFIG.read_text(encoding="utf-8"))
    catalog_bindings = [
        binding
        for binding in deploy_config.get("vectorize", [])
        if binding.get("binding") == "CATALOG_VECTORIZE"
    ]
    if len(catalog_bindings) != 1:
        raise SystemExit("Cannot fingerprint retrieval build; expected one CATALOG_VECTORIZE binding.")
    configured_index = clean(catalog_bindings[0].get("index_name"))
    exported_index = clean(javascript_export(rag_source, "MASTER_VECTORIZE_INDEX"))
    if configured_index != exported_index:
        raise SystemExit(
            "Cannot fingerprint retrieval build; CATALOG_VECTORIZE index does not match MASTER_VECTORIZE_INDEX."
        )

    retrieval_code_sha256 = canonical_sha256(code_files)
    return {
        "retrieval_code_sha256": retrieval_code_sha256,
        "code_files": code_files,
        "bundle_version": javascript_export(rag_source, "MASTER_CATALOG_BUNDLE_VERSION"),
        "default_top_k": javascript_export(rag_source, "MASTER_DEFAULT_TOP_K"),
        "default_chunk_limit": javascript_export(rag_source, "MASTER_DEFAULT_CHUNK_LIMIT"),
        "semantic_score_threshold": javascript_export(
            rag_source, "MASTER_DEFAULT_SEMANTIC_SCORE_THRESHOLD"
        ),
        "embedding": {
            "model": javascript_export(rag_source, "MASTER_EMBEDDING_MODEL"),
            "pooling": javascript_export(rag_source, "MASTER_EMBEDDING_POOLING"),
            "dimensions": javascript_export(rag_source, "MASTER_EMBEDDING_DIMENSIONS"),
        },
        "vectorize": {
            "binding": "CATALOG_VECTORIZE",
            "index_name": configured_index,
            "dimensions": javascript_export(rag_source, "MASTER_EMBEDDING_DIMENSIONS"),
            "metric": javascript_export(rag_source, "MASTER_VECTORIZE_METRIC"),
        },
    }


def retrieval_profile_fingerprint_payload(profile: dict[str, object]) -> dict[str, object]:
    return {
        "fingerprint_schema_version": profile["fingerprint_schema_version"],
        "fixture": {
            "sha256": profile["fixture_sha256"],
            "schema_version": profile["fixture_schema_version"],
            "source_sha256": profile["source_sha256"],
            "required_case_count": profile["fixture_case_count"],
            "raw_case_count": profile["raw_fixture_case_count"],
            "unsupported_case_count": profile["unsupported_case_count"],
        },
        "retrieval_build": profile["retrieval_build"],
    }


def identity_source_fields(record: dict[str, object], extra_fields: list[str] | None = None) -> list[str]:
    fields = [
        "Material Number",
        "Material Description (Global English)",
        "Parent Family Name",
        "Family Name",
    ]
    if is_grounded_value(record.get("Trade Name")):
        fields.append("Trade Name")
    for field in extra_fields or []:
        if field not in fields:
            fields.append(field)
    return fields


def evidence_for(record: dict[str, object], fields: list[str]) -> dict[str, object]:
    return {
        "material_number": record["Material Number"],
        "source_row": record["__source_row"],
        "fields": {field: clean(record.get(field)) for field in fields if is_grounded_value(record.get(field))},
    }


def expected_for_records(
    records: list[dict[str, object]],
    source_fields: list[str],
    *,
    related_material_numbers: list[str] | None = None,
    related_values: list[str] | None = None,
    result_checks: dict[str, object] | None = None,
) -> dict[str, object]:
    ordered_records = sorted(records, key=lambda item: material_sort_key(str(item["Material Number"])))
    return {
        "answerability": "grounded",
        "material_numbers": [str(record["Material Number"]) for record in ordered_records],
        "parent_families": sorted({clean(record["Parent Family Name"]) for record in ordered_records}),
        "families": sorted({clean(record["Family Name"]) for record in ordered_records}),
        "source_fields": source_fields,
        "related_material_numbers": related_material_numbers or [],
        "related_values": related_values or [],
        "evidence": [evidence_for(record, source_fields) for record in ordered_records],
        "result_assertions": result_checks or {},
    }


def query_alias(trade_name: str) -> str:
    value = unicodedata.normalize("NFKD", trade_name).replace("™", "").replace("®", "")
    value = re.sub(r"(?<=[A-Za-z])(?=\d)|(?<=\d)(?=[A-Za-z])", " ", value)
    return re.sub(r"[^A-Za-z0-9]+", " ", value).strip().casefold()


def is_model_like_trade_name(trade_name: str) -> bool:
    value = clean(trade_name)
    compact_value = re.sub(r"[^A-Za-z0-9]", "", value)
    return (
        4 <= len(compact_value) <= 32
        and bool(re.search(r"[A-Za-z]", compact_value))
        and bool(re.search(r"\d", compact_value))
    )


def split_values(value: object) -> list[str]:
    return [part.strip() for part in re.split(r"[;|]", clean(value)) if part.strip()]


def metric_mass_grams(value: object) -> float | None:
    match = re.fullmatch(r"\s*(-?\d+(?:\.\d+)?)\s*(mg|g|kg)\s*", clean(value), re.I)
    if not match:
        return None
    factors = {"mg": 0.001, "g": 1.0, "kg": 1000.0}
    return float(match.group(1)) * factors[match.group(2).casefold()]


def result_assertions(
    *,
    category_name: str,
    material_numbers: list[str],
    numeric_constraints: list[dict[str, object]] | None = None,
) -> dict[str, object]:
    return {
        "catalog_listing_kind": "materials_by_category",
        "category_level": "family",
        "category_name": category_name,
        "exact_listing_material_numbers": sorted(material_numbers, key=material_sort_key),
        "numeric_constraints": numeric_constraints or [],
    }


def load_records(source_path: Path) -> tuple[list[dict[str, object]], int, int]:
    workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    if SOURCE_SHEET not in workbook.sheetnames:
        workbook.close()
        raise SystemExit(f"Workbook does not contain required sheet {SOURCE_SHEET!r}.")

    sheet = workbook[SOURCE_SHEET]
    headers = [clean(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_indexes: dict[str, list[int]] = defaultdict(list)
    for index, header in enumerate(headers):
        header_indexes[header].append(index)

    required_fields = set(IDENTITY_FIELDS)
    required_fields.update(field for field, _, _ in TECHNICAL_CASES)
    required_fields.update(field for field, _, _, _ in RELATIONSHIP_CASES)
    required_fields.update(DOCUMENT_FIELDS)
    required_fields.update({"Alternative Model_#1", "Alternative Model_#2"})
    required_fields.update(
        field
        for _case_id, _material_number, _query, fields in SEMANTIC_DISCOVERY_CASES
        for field in fields
    )

    missing = sorted(field for field in required_fields if field not in header_indexes)
    ambiguous = sorted(field for field in required_fields if len(header_indexes.get(field, [])) > 1)
    if missing or ambiguous:
        workbook.close()
        details = []
        if missing:
            details.append(f"missing fields: {', '.join(missing)}")
        if ambiguous:
            details.append(f"ambiguous fields: {', '.join(ambiguous)}")
        raise SystemExit("Cannot build retrieval fixture; " + "; ".join(details))

    selected_indexes = {field: header_indexes[field][0] for field in sorted(required_fields)}
    records = []
    for source_row, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        record = {field: clean(row[index]) for field, index in selected_indexes.items()}
        record["__source_row"] = source_row
        if is_grounded_value(record["Material Number"]):
            records.append(record)

    row_count = sheet.max_row - 1
    column_count = sheet.max_column
    workbook.close()
    return records, row_count, column_count


def reference_for(
    record: dict[str, object], trade_name_counts: Counter[str], description_counts: Counter[str]
) -> str:
    trade_name = clean(record.get("Trade Name"))
    description = clean(record["Material Description (Global English)"])
    if trade_name and trade_name_counts[normalized(trade_name)] == 1:
        return trade_name
    if description_counts[normalized(description)] == 1:
        return description
    return f"{description} (material {record['Material Number']})"


def record_quality_key(
    record: dict[str, object], trade_name_counts: Counter[str], description_counts: Counter[str]
) -> tuple[int, int, int, str]:
    trade_name = clean(record.get("Trade Name"))
    description = clean(record["Material Description (Global English)"])
    unique_trade_name = bool(trade_name) and trade_name_counts[normalized(trade_name)] == 1
    unique_description = description_counts[normalized(description)] == 1
    material_number = str(record["Material Number"])
    return (
        0 if unique_trade_name else 1,
        0 if unique_description else 1,
        len(material_number),
        material_number,
    )


def build_cases(records: list[dict[str, object]]) -> tuple[list[dict[str, object]], list[dict[str, str]]]:
    by_material = {str(record["Material Number"]): record for record in records}
    by_parent: dict[str, list[dict[str, object]]] = defaultdict(list)
    by_family: dict[str, list[dict[str, object]]] = defaultdict(list)
    for record in records:
        parent = clean(record["Parent Family Name"])
        if parent:
            by_parent[parent].append(record)
        family = clean(record["Family Name"])
        if family:
            by_family[family].append(record)

    trade_name_counts = Counter(
        normalized(record["Trade Name"]) for record in records if is_grounded_value(record.get("Trade Name"))
    )
    description_counts = Counter(normalized(record["Material Description (Global English)"]) for record in records)
    cases: list[dict[str, object]] = []
    exact_parent_manifest: list[dict[str, str]] = []

    for parent in sorted(by_parent, key=str.casefold):
        candidates = sorted(
            by_parent[parent],
            key=lambda record: record_quality_key(record, trade_name_counts, description_counts),
        )
        if parent == "Spare Parts":
            record = by_material["214642"]
        else:
            record = candidates[0]
        fields = identity_source_fields(record)
        material_number = str(record["Material Number"])
        cases.append(
            {
                "id": f"exact-material-parent-{slug(parent)}",
                "category": "exact_material",
                "subcategory": "parent_family_lookup",
                "coverage_key": parent,
                "query": f"What catalog item is OHAUS material number {material_number}?",
                "expected": expected_for_records([record], fields),
            }
        )
        exact_parent_manifest.append({"parent_family": parent, "material_number": material_number})

    exact_parent_materials = {entry["material_number"] for entry in exact_parent_manifest}
    for material_number in SIX_DIGIT_MATERIALS:
        if material_number in exact_parent_materials:
            continue
        record = by_material.get(material_number)
        if record is None:
            raise SystemExit(f"Required six-digit material {material_number} is missing from {SOURCE_SHEET}.")
        fields = identity_source_fields(record)
        cases.append(
            {
                "id": f"exact-material-six-digit-{material_number}",
                "category": "exact_material",
                "subcategory": "six_digit_material_lookup",
                "coverage_key": material_number,
                "query": f"Identify OHAUS material number {material_number}.",
                "expected": expected_for_records([record], fields),
            }
        )

    for field, alias in [("Alternative Model_#1", "9123"), ("Alternative Model_#2", "713")]:
        alias_records = sorted(
            [record for record in records if clean(record.get(field)) == alias],
            key=lambda record: material_sort_key(str(record["Material Number"])),
        )
        if not alias_records:
            raise SystemExit(f"No rows found for required alternative model alias {alias!r} in {field!r}.")
        fields = ["Material Number", "Material Description (Global English)", "Parent Family Name", "Family Name", field]
        cases.append(
            {
                "id": f"model-alias-alternative-{alias}",
                "category": "model_alias",
                "subcategory": "alternative_model",
                "coverage_key": field,
                "query": f"Which catalog materials use the alternative model number {alias}?",
                "expected": expected_for_records(alias_records, fields),
            }
        )

    parent_counts = Counter(clean(record["Parent Family Name"]) for record in records if clean(record["Parent Family Name"]))
    unique_trade_candidates: dict[str, list[dict[str, object]]] = defaultdict(list)
    for record in records:
        trade_name = clean(record.get("Trade Name"))
        parent = clean(record["Parent Family Name"])
        if (
            trade_name
            and parent
            and is_model_like_trade_name(trade_name)
            and trade_name_counts[normalized(trade_name)] == 1
        ):
            unique_trade_candidates[parent].append(record)

    selected_trade_records = []
    for parent in sorted(unique_trade_candidates, key=lambda value: (-parent_counts[value], value.casefold())):
        candidates = sorted(
            unique_trade_candidates[parent],
            key=lambda record: (len(clean(record["Trade Name"])), normalized(record["Trade Name"]), material_sort_key(str(record["Material Number"]))),
        )
        selected_trade_records.append(candidates[0])
        if len(selected_trade_records) == 13:
            break
    if len(selected_trade_records) != 13:
        raise SystemExit("Could not select 13 unique Trade Name alias cases.")

    for sequence, record in enumerate(selected_trade_records, start=1):
        trade_name = clean(record["Trade Name"])
        fields = identity_source_fields(record)
        cases.append(
            {
                "id": f"model-alias-trade-name-{sequence:02d}-{record['Material Number']}",
                "category": "model_alias",
                "subcategory": "trade_name_normalization",
                "coverage_key": clean(record["Parent Family Name"]),
                "query": f"Which OHAUS material corresponds to the model alias {query_alias(trade_name)}?",
                "expected": expected_for_records([record], fields),
            }
        )

    used_technical_materials: set[str] = set()
    for field, subcategory, question_template in TECHNICAL_CASES:
        candidates = [record for record in records if is_grounded_value(record.get(field))]
        candidates.sort(key=lambda record: record_quality_key(record, trade_name_counts, description_counts))
        unused_candidates = [record for record in candidates if str(record["Material Number"]) not in used_technical_materials]
        record = (unused_candidates or candidates)[0]
        used_technical_materials.add(str(record["Material Number"]))
        fields = identity_source_fields(record, [field])
        reference = reference_for(record, trade_name_counts, description_counts)
        cases.append(
            {
                "id": f"technical-{subcategory}-{record['Material Number']}",
                "category": "technical_field",
                "subcategory": subcategory,
                "coverage_key": field,
                "query": question_template.format(reference=reference),
                "expected": expected_for_records([record], fields),
            }
        )

    for field, subcategory, question_template, answer_kind in RELATIONSHIP_CASES:
        candidates = [record for record in records if is_grounded_value(record.get(field))]
        candidates.sort(
            key=lambda record: (
                len(split_values(record[field])),
                record_quality_key(record, trade_name_counts, description_counts),
            )
        )
        record = candidates[0]
        fields = identity_source_fields(record, [field])
        reference = reference_for(record, trade_name_counts, description_counts)
        values = split_values(record[field])
        related_material_numbers = values if answer_kind == "material_ids" else []
        related_values = values if answer_kind == "values" else []
        cases.append(
            {
                "id": f"relationship-{subcategory}-{record['Material Number']}",
                "category": "relationship",
                "subcategory": subcategory,
                "coverage_key": field,
                "query": question_template.format(reference=reference),
                "expected": expected_for_records(
                    [record],
                    fields,
                    related_material_numbers=related_material_numbers,
                    related_values=related_values,
                ),
            }
        )

    for field in DOCUMENT_FIELDS:
        candidates = [record for record in records if is_grounded_value(record.get(field))]
        candidates.sort(key=lambda record: record_quality_key(record, trade_name_counts, description_counts))
        record = candidates[0]
        fields = identity_source_fields(record, [field])
        reference = reference_for(record, trade_name_counts, description_counts)
        cases.append(
            {
                "id": f"document-{slug(field)}-{record['Material Number']}",
                "category": "document_link",
                "subcategory": "specific_document_field",
                "coverage_key": field,
                "query": f"What {field.removeprefix('EN ').rsplit(' ', 1)[0].lower()} link is listed for {reference}?",
                "expected": expected_for_records([record], fields),
            }
        )

    combined_document_candidates = [
        record for record in records if any(is_grounded_value(record.get(field)) for field in DOCUMENT_FIELDS)
    ]
    combined_document_candidates.sort(
        key=lambda record: (
            -sum(is_grounded_value(record.get(field)) for field in DOCUMENT_FIELDS),
            record_quality_key(record, trade_name_counts, description_counts),
        )
    )
    combined_record = combined_document_candidates[0]
    combined_fields = [field for field in DOCUMENT_FIELDS if is_grounded_value(combined_record.get(field))]
    fields = identity_source_fields(combined_record, combined_fields)
    reference = reference_for(combined_record, trade_name_counts, description_counts)
    cases.append(
        {
            "id": f"document-all-english-{combined_record['Material Number']}",
            "category": "document_link",
            "subcategory": "all_english_documents",
            "coverage_key": "all_english_documents",
            "query": f"Which English data sheets, user guides, and manuals are listed for {reference}?",
            "expected": expected_for_records([combined_record], fields),
        }
    )

    for case_id, material_number, query, discovery_fields in SEMANTIC_DISCOVERY_CASES:
        record = by_material.get(material_number)
        if record is None:
            raise SystemExit(f"Semantic discovery anchor {material_number} is missing from {SOURCE_SHEET}.")
        if any(not is_grounded_value(record.get(field)) for field in discovery_fields):
            raise SystemExit(f"Semantic discovery anchor {material_number} has incomplete discovery evidence.")
        fields = identity_source_fields(record, discovery_fields)
        expected = expected_for_records([record], fields)
        expected["result_assertions"] = {"semantic_material_numbers": [material_number]}
        cases.append(
            {
                "id": f"semantic-discovery-{case_id}-{material_number}",
                "category": "semantic_discovery",
                "subcategory": case_id,
                "coverage_key": material_number,
                "query": query,
                "expected": expected,
            }
        )

    compass_records = sorted(by_family["Compass™ CR"], key=lambda item: material_sort_key(str(item["Material Number"])))
    compass_materials = [str(record["Material Number"]) for record in compass_records]
    cases.append(
        {
            "id": "category-family-compass-cr",
            "category": "category_filter",
            "subcategory": "complete_family_listing",
            "coverage_key": "Compass™ CR",
            "query": "Show every Compass CR model in the catalog.",
            "expected": expected_for_records(
                compass_records,
                ["Material Number", "Material Description (Global English)", "Parent Family Name", "Family Name", "Trade Name"],
                result_checks=result_assertions(
                    category_name="Compass™ CR",
                    material_numbers=compass_materials,
                ),
            ),
        }
    )

    numeric_case_specs = [
        {
            "id": "numeric-compass-capacity-exact-620g",
            "family": "Compass™ CR",
            "query": "Which Compass CR models have a maximum capacity of exactly 620 g?",
            "capacity_min_g": 620.0,
            "capacity_max_g": 620.0,
            "readability_g": None,
            "constraints": [{"field": "capacity", "comparator": "exact", "value": 620, "unit": "g"}],
        },
        {
            "id": "numeric-scout-capacity-range-400g-700g",
            "family": "Scout™ STX",
            "query": "Which Scout STX models have a maximum capacity between 400 g and 700 g?",
            "capacity_min_g": 400.0,
            "capacity_max_g": 700.0,
            "readability_g": None,
            "constraints": [
                {"field": "capacity", "comparator": "at_least", "value": 400, "unit": "g"},
                {"field": "capacity", "comparator": "at_most", "value": 700, "unit": "g"},
            ],
        },
        {
            "id": "numeric-scout-capacity-and-readability",
            "family": "Scout™ STX",
            "query": "Which Scout STX models have capacity between 100 g and 250 g and readability of exactly 0.001 g?",
            "capacity_min_g": 100.0,
            "capacity_max_g": 250.0,
            "readability_g": 0.001,
            "constraints": [
                {"field": "capacity", "comparator": "at_least", "value": 100, "unit": "g"},
                {"field": "capacity", "comparator": "at_most", "value": 250, "unit": "g"},
                {"field": "readability", "comparator": "exact", "value": 0.001, "unit": "g"},
            ],
        },
    ]
    for spec in numeric_case_specs:
        matches = []
        for record in by_family[str(spec["family"])]:
            capacity = metric_mass_grams(record.get("Maximum Capacity {metric}"))
            readability = metric_mass_grams(record.get("Readability {metric}"))
            if capacity is None or not float(spec["capacity_min_g"]) <= capacity <= float(spec["capacity_max_g"]):
                continue
            if spec["readability_g"] is not None and readability != float(spec["readability_g"]):
                continue
            matches.append(record)
        if not matches:
            raise SystemExit(f"Numeric evaluation case {spec['id']} has no workbook-grounded matches.")
        technical_fields = ["Maximum Capacity {metric}"]
        if spec["readability_g"] is not None:
            technical_fields.append("Readability {metric}")
        fields = ["Material Number", "Material Description (Global English)", "Parent Family Name", "Family Name", "Trade Name", *technical_fields]
        material_numbers = [str(record["Material Number"]) for record in matches]
        cases.append(
            {
                "id": str(spec["id"]),
                "category": "numeric_filter",
                "subcategory": "range" if len(spec["constraints"]) > 1 else "exact",
                "coverage_key": str(spec["family"]),
                "query": str(spec["query"]),
                "expected": expected_for_records(
                    matches,
                    fields,
                    result_checks=result_assertions(
                        category_name=str(spec["family"]),
                        material_numbers=material_numbers,
                        numeric_constraints=list(spec["constraints"]),
                    ),
                ),
            }
        )

    cases.append(
        {
            "id": "no-results-compass-capacity-range-700g-800g",
            "category": "no_results",
            "subcategory": "empty_category_numeric_range",
            "coverage_key": "Compass™ CR",
            "query": "Which Compass CR models have a maximum capacity between 700 g and 800 g?",
            "expected": {
                "answerability": "no_results",
                "material_numbers": [],
                "parent_families": ["Portable Balances"],
                "families": ["Compass™ CR"],
                "source_fields": [],
                "related_material_numbers": [],
                "related_values": [],
                "evidence": [],
                "required_behavior": "Return an empty filtered family listing; do not substitute non-matching products.",
                "result_assertions": result_assertions(
                    category_name="Compass™ CR",
                    material_numbers=[],
                    numeric_constraints=[
                        {"field": "capacity", "comparator": "at_least", "value": 700, "unit": "g"},
                        {"field": "capacity", "comparator": "at_most", "value": 800, "unit": "g"},
                    ],
                ),
            },
        }
    )

    unsupported_anchors = [
        by_material[entry["material_number"]]
        for entry in exact_parent_manifest[: len(UNSUPPORTED_CASES)]
    ]
    for (subcategory, question_template), record in zip(UNSUPPORTED_CASES, unsupported_anchors):
        material_number = str(record["Material Number"])
        cases.append(
            {
                "id": f"unsupported-{subcategory}-{material_number}",
                "category": "unsupported_live_data",
                "subcategory": subcategory,
                "coverage_key": subcategory,
                "query": question_template.format(material_number=material_number),
                "anchor_material_number": material_number,
                "expected": {
                    "answerability": "unsupported_live_data",
                    "material_numbers": [],
                    "parent_families": [],
                    "families": [],
                    "source_fields": [],
                    "related_material_numbers": [],
                    "related_values": [],
                    "evidence": [],
                    "result_assertions": {},
                    "required_behavior": "State that the static master workbook does not provide the requested live value and do not invent one.",
                },
            }
        )

    return cases, exact_parent_manifest


def validate_fixture(cases: list[dict[str, object]], exact_parent_manifest: list[dict[str, str]]) -> None:
    ids = [str(case["id"]) for case in cases]
    if len(ids) != len(set(ids)):
        raise SystemExit("Generated duplicate evaluation case IDs.")
    if not 80 <= len(cases) <= 120:
        raise SystemExit(f"Expected 80-120 evaluation cases, generated {len(cases)}.")
    if len(exact_parent_manifest) != 46:
        raise SystemExit(f"Expected 46 parent-family lookup cases, generated {len(exact_parent_manifest)}.")

    grounded = [case for case in cases if case["expected"]["answerability"] == "grounded"]
    for case in grounded:
        expected = case["expected"]
        if not expected["material_numbers"]:
            raise SystemExit(f"Grounded case {case['id']} has no expected material numbers.")
        if not expected["source_fields"] or not expected["evidence"]:
            raise SystemExit(f"Grounded case {case['id']} has incomplete source evidence.")
        if not expected["parent_families"] or not expected["families"]:
            raise SystemExit(f"Grounded case {case['id']} has incomplete category evidence.")
    required = [case for case in cases if case["expected"]["answerability"] != "unsupported_live_data"]
    no_results = [case for case in required if case["expected"]["answerability"] == "no_results"]
    if not no_results:
        raise SystemExit("Generated fixture must contain a required no-results case.")
    for case in no_results:
        assertions = case["expected"].get("result_assertions", {})
        if assertions.get("exact_listing_material_numbers") != []:
            raise SystemExit(f"No-results case {case['id']} must require an empty exact listing.")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--profile-output", type=Path, default=DEFAULT_PROFILE_OUTPUT)
    args = parser.parse_args()

    source_path = args.source.resolve()
    output_path = args.output.resolve()
    profile_output_path = args.profile_output.resolve()
    if not source_path.is_file():
        raise SystemExit(f"Workbook not found: {source_path}")

    records, source_rows, source_columns = load_records(source_path)
    cases, exact_parent_manifest = build_cases(records)
    validate_fixture(cases, exact_parent_manifest)

    categories = Counter(str(case["category"]) for case in cases)
    grounded_case_count = sum(case["expected"]["answerability"] == "grounded" for case in cases)
    required_case_count = sum(case["expected"]["answerability"] != "unsupported_live_data" for case in cases)
    grounded_material_numbers = sorted(
        {
            material
            for case in cases
            if case["category"] != "unsupported_live_data"
            for material in case["expected"]["material_numbers"]
        },
        key=material_sort_key,
    )
    parent_families = [entry["parent_family"] for entry in exact_parent_manifest]
    exact_parent_manifest_sha256 = canonical_sha256(exact_parent_manifest)

    fixture = {
        "schema_version": "2.0.0",
        "source": {
            "file": source_path.name,
            "sheet": SOURCE_SHEET,
            "sha256": file_sha256(source_path),
            "rows": source_rows,
            "columns": source_columns,
            "selection_rule": "Use Product_Catalog_AI only; exclude Raw_Data and blank Parent Family Name values.",
        },
        "manifest": {
            "case_count": len(cases),
            "grounded_case_count": grounded_case_count,
            "required_case_count": required_case_count,
            "unsupported_case_count": categories["unsupported_live_data"],
            "categories": dict(sorted(categories.items())),
            "parent_family_count": len(parent_families),
            "parent_families": parent_families,
            "exact_parent_family_materials": exact_parent_manifest,
            "exact_parent_family_materials_sha256": exact_parent_manifest_sha256,
            "six_digit_material_numbers": SIX_DIGIT_MATERIALS,
            "grounded_material_numbers": grounded_material_numbers,
            "technical_fields": [field for field, _, _ in TECHNICAL_CASES],
            "relationship_fields": [field for field, _, _, _ in RELATIONSHIP_CASES],
            "document_fields": DOCUMENT_FIELDS,
        },
        "cases": cases,
    }

    output_text = json.dumps(fixture, ensure_ascii=False, indent=2) + "\n"
    fixture_sha256 = hashlib.sha256(output_text.encode("utf-8")).hexdigest()
    retrieval_build = retrieval_build_profile()
    profile = {
        "profile_version": PROFILE_VERSION,
        "fingerprint_schema_version": FINGERPRINT_SCHEMA_VERSION,
        "fixture_file": output_path.name,
        "fixture_sha256": fixture_sha256,
        "fixture_schema_version": fixture["schema_version"],
        "source_sha256": fixture["source"]["sha256"],
        "fixture_case_count": fixture["manifest"]["required_case_count"],
        "raw_fixture_case_count": fixture["manifest"]["case_count"],
        "unsupported_case_count": fixture["manifest"]["unsupported_case_count"],
        "retrieval_build": retrieval_build,
    }
    profile["retrieval_profile_sha256"] = canonical_sha256(retrieval_profile_fingerprint_payload(profile))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    profile_output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(output_text, encoding="utf-8")
    profile_output_path.write_text(json.dumps(profile, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(
        json.dumps(
            {
                "output": str(output_path),
                "profile_output": str(profile_output_path),
                "fixture_sha256": fixture_sha256,
                "retrieval_profile_sha256": profile["retrieval_profile_sha256"],
                "cases": len(cases),
                "categories": dict(sorted(categories.items())),
                "parent_families": len(parent_families),
                "source_sha256": fixture["source"]["sha256"],
                "parent_material_manifest_sha256": exact_parent_manifest_sha256,
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
