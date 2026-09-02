#!/usr/bin/env python3
"""Build a staged, source-traceable D1 catalog from the MMMDF workbook.

This is an offline maintenance command. It never activates a catalog. It emits
sharded D1 SQL plus QA, record, chunk, and Vectorize-seed artifacts. The final
activation statement is deliberately kept in a separate generated SQL file.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# Both values participate in every future catalog version ID. Bump the schema
# version for storage-contract changes and the generator version whenever an
# importer change can alter generated artifacts for an unchanged workbook.
CATALOG_SCHEMA_VERSION = "master-catalog-v1"
CATALOG_GENERATOR_VERSION = "1.2.0"
DEFAULT_SHEET = "Product_Catalog_AI"
MAX_CHUNK_CHARACTERS = 1_450
SQL_SHARD_BYTES = 8 * 1024 * 1024
SQL_BATCH_ROWS = 100
MAX_INSERT_STATEMENT_BYTES = 80 * 1024

RELATIONSHIP_PREFIX = "Relationship / "
DOCUMENT_HEADER = re.compile(r"^EN (Data Sheets|User Guide|Manuals) (\d+)$", re.I)
ALTERNATIVE_MODEL_HEADER = re.compile(r"^Alternative Model(?:_#\d+)?$", re.I)
URL_PATTERN = re.compile(r"https?://[^\s;,|]+", re.I)
MEASUREMENT_PATTERN = re.compile(
    r"^\s*(?:[<>≤≥~=±]+\s*)?([+-]?(?:\d+(?:,\d{3})*|\d*)(?:\.\d+)?)\s*([^\d]*)\s*$"
)
IDENTIFIER_FIELD_MARKERS = (
    "material_number", "commodity_code", "product_hierarchy", "service_hierarchy",
    "model", "trade_name", "url", "document", "phone", "postal", "zip",
)
SEMANTIC_EXCLUDED_HEADER_MARKERS = (
    "image", "sales org", "delivering plant", "procurement type", "commodity code",
    "minimumorderquantity", "minimum order", "order note",
)
SEMANTIC_EXCLUDED_SEARCH_SEGMENTS = (
    "sales org", "plant", "procurement", "commodity", "minimum order", "order note",
)
SEMANTIC_PLACEHOLDERS = {"n a", "na", "not applicable", "none"}

FIELD_GROUPS = (
    ("identity", ("material", "description", "parent_family", "family_name", "trade_name", "hierarchy")),
    ("discovery", ("ai_search", "ai_summary", "tag_line", "benefit", "key_feature", "design_feature")),
    ("applications", ("application", "product_will_be_used", "typical_area", "market_world", "operation")),
    ("commercial", ("sales_org", "plant", "order", "procurement", "commodity", "origin", "base_unit", "delivery_unit")),
    ("physical", ("dimension", "height", "length", "width", "diameter", "weight", "construction", "material_")),
    ("connectivity", ("communication", "interface", "usb", "ethernet", "wireless", "bluetooth", "rs232")),
    ("environment", ("environment", "temperature", "humidity", "altitude", "protection", "ip_rating")),
    ("compliance", ("legal", "approval", "certificate", "class", "standard", "accuracy")),
    ("performance", ("capacity", "readability", "resolution", "speed", "range", "time", "power", "battery", "voltage", "frequency")),
)


@dataclass(frozen=True)
class Header:
    ordinal: int
    column: str
    source_header: str
    field_key: str


@dataclass
class Record:
    material_number: str
    product_name: str
    parent_family: str
    family: str
    trade_name: str
    ai_summary: str
    ai_search_index: str
    source_row: int
    fields: dict[str, Any]
    record_json: str
    record_sha256: str


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_text(value: str) -> str:
    return sha256_bytes(value.encode("utf-8"))


def stable_id(prefix: str, *parts: object) -> str:
    digest = sha256_text("\x1f".join(str(part) for part in parts))[:40]
    return f"{prefix}_{digest}"


def normalize_space(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def normalize_alias(value: object) -> str:
    text = unicodedata.normalize("NFKD", normalize_space(value))
    text = "".join(character for character in text if not unicodedata.combining(character))
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def identifier_alias_normalizations(alias_type: str, value: object) -> list[str]:
    """Return deterministic exact-match forms for model-like identifiers.

    The source-normalized form is always first. Identifier fields also receive
    a compact form and a form split at every alpha/digit boundary, allowing a
    query such as ``CR221`` to match source data stored as ``CR 221`` (and vice
    versa) through the indexed ``master_aliases.normalized_alias`` column.
    """
    normalized = normalize_alias(value)
    if not normalized:
        return []
    variants = [normalized]
    if alias_type not in {"material_number", "trade_name", "alternative_model"}:
        return variants
    compact = normalized.replace(" ", "")
    if not (re.search(r"[a-z]", compact) and re.search(r"\d", compact)):
        return variants
    spaced = re.sub(r"([a-z])([0-9])", r"\1 \2", compact)
    spaced = re.sub(r"([0-9])([a-z])", r"\1 \2", spaced)
    return list(dict.fromkeys((normalized, compact, spaced)))


def split_alias_values(value: object) -> list[str]:
    """Split list-like alternative-model cells without breaking model slashes."""
    return list(dict.fromkeys(
        token.strip() for token in re.split(r"[;,|]", value_text(value)) if token.strip()
    ))


def snake(value: object) -> str:
    text = unicodedata.normalize("NFKD", normalize_space(value)).lower()
    text = "".join(character for character in text if not unicodedata.combining(character))
    text = text.replace("±", " plus_minus ").replace("ø", " diameter ")
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def json_value(value: object) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return int(value) if value.is_integer() else value
    if isinstance(value, str):
        return value.strip()
    return str(value)


def value_text(value: object) -> str:
    converted = json_value(value)
    if converted is None:
        return ""
    if isinstance(converted, bool):
        return "true" if converted else "false"
    return normalize_space(converted)


def populated(value: object) -> bool:
    converted = json_value(value)
    return converted is not None and converted != ""


def disambiguate_headers(values: Iterable[object]) -> list[Header]:
    counts: Counter[str] = Counter()
    headers: list[Header] = []
    for ordinal, value in enumerate(values, start=1):
        column = get_column_letter(ordinal)
        source_header = normalize_space(value) or f"Unnamed Column {column}"
        base = snake(source_header) or f"column_{column.lower()}"
        counts[base] += 1
        field_key = base if counts[base] == 1 else f"{base}__{counts[base]}"
        headers.append(Header(ordinal, column, source_header, field_key))
    return headers


def first_field_key(headers: list[Header], source_header: str) -> str | None:
    return next((header.field_key for header in headers if header.source_header == source_header), None)


def record_field(record_fields: dict[str, Any], headers: list[Header], source_header: str) -> str:
    key = first_field_key(headers, source_header)
    return value_text(record_fields.get(key)) if key else ""


def parse_numeric(field_key: str, value: object) -> tuple[float | None, str | None]:
    if any(marker in field_key for marker in IDENTIFIER_FIELD_MARKERS):
        return None, None
    if isinstance(value, bool):
        return None, None
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value), None
    text = value_text(value)
    if not text or len(text) > 80:
        return None, None
    match = MEASUREMENT_PATTERN.fullmatch(text)
    if not match or not match.group(1):
        return None, None
    try:
        number = float(match.group(1).replace(",", ""))
    except ValueError:
        return None, None
    unit = normalize_space(match.group(2)).strip("()[]{}") or None
    if unit and len(unit) > 32:
        return None, None
    return number, unit


def canonical_measurement(number: float | None, unit: str | None, field_key: str) -> tuple[float | None, str | None]:
    if number is None or not unit:
        return None, None
    normalized = normalize_space(unit).lower().replace("μ", "µ").replace("°", "")
    normalized = normalized.rstrip(".")
    conversions: dict[str, tuple[float, float, str]] = {
        "kg": (1_000, 0, "g"), "kilogram": (1_000, 0, "g"), "kilograms": (1_000, 0, "g"),
        "g": (1, 0, "g"), "gram": (1, 0, "g"), "grams": (1, 0, "g"),
        "mg": (0.001, 0, "g"), "milligram": (0.001, 0, "g"), "milligrams": (0.001, 0, "g"),
        "µg": (0.000001, 0, "g"), "ug": (0.000001, 0, "g"),
        "lb": (453.59237, 0, "g"), "lbs": (453.59237, 0, "g"), "pound": (453.59237, 0, "g"), "pounds": (453.59237, 0, "g"),
        "oz": (28.349523125, 0, "g"), "ounce": (28.349523125, 0, "g"), "ounces": (28.349523125, 0, "g"),
        "mm": (1, 0, "mm"), "millimeter": (1, 0, "mm"), "millimeters": (1, 0, "mm"),
        "cm": (10, 0, "mm"), "centimeter": (10, 0, "mm"), "centimeters": (10, 0, "mm"),
        "m": (1_000, 0, "mm"), "meter": (1_000, 0, "mm"), "meters": (1_000, 0, "mm"),
        "in": (25.4, 0, "mm"), "inch": (25.4, 0, "mm"), "inches": (25.4, 0, "mm"),
        "ft": (304.8, 0, "mm"), "foot": (304.8, 0, "mm"), "feet": (304.8, 0, "mm"),
        "ml": (1, 0, "mL"), "milliliter": (1, 0, "mL"), "milliliters": (1, 0, "mL"),
        "l": (1_000, 0, "mL"), "liter": (1_000, 0, "mL"), "liters": (1_000, 0, "mL"),
        "µl": (0.001, 0, "mL"), "ul": (0.001, 0, "mL"),
        "ms": (0.001, 0, "s"), "millisecond": (0.001, 0, "s"), "milliseconds": (0.001, 0, "s"),
        "s": (1, 0, "s"), "sec": (1, 0, "s"), "second": (1, 0, "s"), "seconds": (1, 0, "s"),
        "min": (60, 0, "s"), "minute": (60, 0, "s"), "minutes": (60, 0, "s"),
        "h": (3_600, 0, "s"), "hr": (3_600, 0, "s"), "hour": (3_600, 0, "s"), "hours": (3_600, 0, "s"),
    }
    if normalized in {"c", "celsius"} and any(marker in field_key for marker in ("temperature", "environment")):
        return number, "°C"
    if normalized in {"f", "fahrenheit"} and any(marker in field_key for marker in ("temperature", "environment")):
        return (number - 32) * 5 / 9, "°C"
    if normalized in {"k", "kelvin"} and any(marker in field_key for marker in ("temperature", "environment")):
        return number - 273.15, "°C"
    if normalized == "m" and any(marker in field_key for marker in ("time", "duration")):
        return number * 60, "s"
    conversion = conversions.get(normalized)
    if not conversion:
        return None, None
    multiplier, offset, canonical_unit = conversion
    return number * multiplier + offset, canonical_unit


def field_group(field_key: str) -> str:
    if field_key.startswith("relationship_"):
        return "relationships"
    if DOCUMENT_HEADER.match(field_key.replace("_", " ")):
        return "documents"
    for group, markers in FIELD_GROUPS:
        if any(marker in field_key for marker in markers):
            return group
    return "additional"


def split_relationships(value: object) -> list[str]:
    return list(dict.fromkeys(
        token.strip() for token in re.split(r"[;,|]", value_text(value)) if token.strip()
    ))


def document_type(source_header: str) -> str | None:
    match = DOCUMENT_HEADER.match(source_header)
    if not match:
        return None
    return {
        "data sheets": "data_sheet",
        "user guide": "user_guide",
        "manuals": "manual",
    }[match.group(1).lower()]


def document_urls(value: object) -> list[str]:
    text = value_text(value)
    matches = [match.rstrip(".)]") for match in URL_PATTERN.findall(text)]
    return list(dict.fromkeys(matches or ([text] if text else [])))


def semantic_value(source_header: str, value: object) -> str:
    """Remove internal-only derived segments while retaining the source field itself."""
    text = value_text(value)
    if normalize_alias(text) in SEMANTIC_PLACEHOLDERS:
        return ""
    if source_header != "AI_Search_Index":
        return text
    segments = [normalize_space(segment) for segment in text.split("|")]
    return " | ".join(
        segment for segment in segments
        if segment and not any(
            normalize_alias(segment).startswith(marker) for marker in SEMANTIC_EXCLUDED_SEARCH_SEGMENTS
        )
    )


def sql_literal(value: object) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (int, float)):
        if not math.isfinite(float(value)):
            return "NULL"
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


class SqlShardWriter:
    def __init__(self, output_dir: Path, version_id: str):
        self.output_dir = output_dir
        self.version_id = version_id
        self.index = 0
        self.handle = None
        self.current_bytes = 0
        self.paths: list[Path] = []
        self.maximum_statement_bytes = 0

    def _open(self) -> None:
        self.index += 1
        path = self.output_dir / f"master-catalog-stage-{self.index:04d}.sql"
        self.handle = path.open("w", encoding="utf-8", newline="\n")
        self.paths.append(path)
        header = (
            "-- Generated staged master-catalog import. Apply migration first.\n"
            f"-- Catalog version: {self.version_id}\n"
            "PRAGMA foreign_keys = ON;\n"
        )
        self.handle.write(header)
        self.current_bytes = len(header.encode("utf-8"))

    def append(self, statement: str) -> None:
        payload = statement.rstrip() + "\n"
        size = len(payload.encode("utf-8"))
        self.maximum_statement_bytes = max(self.maximum_statement_bytes, size)
        if self.handle is None or (self.current_bytes + size > SQL_SHARD_BYTES and self.current_bytes > 200):
            if self.handle is not None:
                self.handle.close()
            self._open()
        self.handle.write(payload)
        self.current_bytes += size

    def close(self) -> list[Path]:
        if self.handle is not None:
            self.handle.close()
            self.handle = None
        return self.paths


class BatchedInsert:
    def __init__(self, writer: SqlShardWriter, table: str, columns: list[str]):
        self.writer = writer
        self.table = table
        self.columns = columns
        self.rows: list[str] = []
        self.prefix = f"INSERT OR IGNORE INTO {self.table}({','.join(self.columns)}) VALUES\n"
        self.encoded_bytes = len(self.prefix.encode("utf-8")) + 2

    def fits_single(self, row: tuple[object, ...]) -> bool:
        encoded_row = "(" + ",".join(sql_literal(value) for value in row) + ")"
        return len(self.prefix.encode("utf-8")) + len(encoded_row.encode("utf-8")) + 2 <= MAX_INSERT_STATEMENT_BYTES

    def add(self, row: tuple[object, ...]) -> None:
        encoded_row = "(" + ",".join(sql_literal(value) for value in row) + ")"
        row_bytes = len(encoded_row.encode("utf-8")) + (2 if self.rows else 0)
        if self.rows and self.encoded_bytes + row_bytes > MAX_INSERT_STATEMENT_BYTES:
            self.flush()
            row_bytes = len(encoded_row.encode("utf-8"))
        if self.encoded_bytes + row_bytes > MAX_INSERT_STATEMENT_BYTES:
            raise ValueError(
                f"One {self.table} row exceeds the conservative SQL statement limit: "
                f"{self.encoded_bytes + row_bytes} bytes"
            )
        self.rows.append(encoded_row)
        self.encoded_bytes += row_bytes
        if len(self.rows) >= SQL_BATCH_ROWS:
            self.flush()

    def flush(self) -> None:
        if not self.rows:
            return
        self.writer.append(self.prefix + ",\n".join(self.rows) + ";")
        self.rows.clear()
        self.encoded_bytes = len(self.prefix.encode("utf-8")) + 2


def chunk_prefix(record: Record, kind: str) -> str:
    name = record.trade_name or record.product_name or record.material_number
    parts = [
        f"Category: {record.parent_family or 'Uncategorized'}",
        f"Family: {record.family or 'Uncategorized'}",
        f"Material number: {record.material_number}",
        f"Product: {name}",
        f"Section: {kind.replace('_', ' ').title()}",
    ]
    return "\n".join(parts) + "\n"


def split_long_line(line: str, maximum: int) -> list[str]:
    if len(line) <= maximum:
        return [line]
    words = line.split()
    if len(words) <= 1:
        return [line[index:index + maximum] for index in range(0, len(line), maximum)]
    chunks: list[str] = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if len(candidate) > maximum and current:
            chunks.append(current)
            current = word
        else:
            current = candidate
    if current:
        chunks.append(current)
    return chunks


def pack_chunks(record: Record, version_id: str, kind: str, lines: list[tuple[str, str]]) -> list[dict[str, Any]]:
    prefix = chunk_prefix(record, kind)
    available = MAX_CHUNK_CHARACTERS - len(prefix)
    if available < 200:
        raise ValueError(f"Chunk prefix unexpectedly long for material {record.material_number}")
    expanded: list[tuple[str, str]] = []
    for field_key, line in lines:
        for segment in split_long_line(line, available):
            expanded.append((field_key, segment))

    packed: list[dict[str, Any]] = []
    current_lines: list[str] = []
    current_keys: list[str] = []
    for field_key, line in expanded:
        candidate = "\n".join([*current_lines, line])
        if current_lines and len(prefix) + len(candidate) > MAX_CHUNK_CHARACTERS:
            packed.append({"lines": current_lines, "field_keys": list(dict.fromkeys(current_keys))})
            current_lines = [line]
            current_keys = [field_key]
        else:
            current_lines.append(line)
            current_keys.append(field_key)
    if current_lines:
        packed.append({"lines": current_lines, "field_keys": list(dict.fromkeys(current_keys))})

    output: list[dict[str, Any]] = []
    title_name = record.trade_name or record.product_name or record.material_number
    for ordinal, item in enumerate(packed, start=1):
        content = prefix + "\n".join(item["lines"])
        if len(content) > MAX_CHUNK_CHARACTERS:
            raise ValueError(f"Chunk exceeds limit for material {record.material_number}: {len(content)}")
        chunk_id = stable_id("mc", version_id, record.record_sha256, kind, ordinal, content)
        output.append({
            "chunk_id": chunk_id,
            "material_number": record.material_number,
            "parent_family": record.parent_family,
            "family": record.family,
            "chunk_kind": kind,
            "chunk_ordinal": ordinal,
            "title": f"{title_name} — {kind.replace('_', ' ').title()}",
            "content": content,
            "content_sha256": sha256_text(content),
            "field_keys": item["field_keys"],
        })
    return output


def record_chunks(record: Record, headers: list[Header], material_numbers: set[str], version_id: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    headers_by_key = {header.field_key: header for header in headers}
    grouped: dict[str, list[tuple[str, str]]] = defaultdict(list)
    relationships: list[dict[str, Any]] = []
    documents: list[dict[str, Any]] = []
    seen_relationships: set[tuple[str, str]] = set()
    seen_documents: set[tuple[str, str]] = set()

    for field_key, value in record.fields.items():
        header = headers_by_key[field_key]
        if header.source_header.startswith(RELATIONSHIP_PREFIX):
            relation_type = snake(header.source_header.removeprefix(RELATIONSHIP_PREFIX))
            for ordinal, target in enumerate(split_relationships(value), start=1):
                relation_key = (relation_type, target)
                if relation_key in seen_relationships:
                    continue
                seen_relationships.add(relation_key)
                relationships.append({
                    # IDs are globally unique in D1. Include the immutable catalog
                    # version so an unchanged relationship can coexist in a later
                    # staged version instead of being skipped by INSERT OR IGNORE.
                    "relationship_id": stable_id("rel", version_id, record.record_sha256, relation_type, target),
                    "source_material_number": record.material_number,
                    "relationship_type": relation_type,
                    "target_material_number": target,
                    "target_resolved": target in material_numbers,
                    "source_field": field_key,
                    "source_ordinal": ordinal,
                })
            continue

        kind = document_type(header.source_header)
        if kind:
            for ordinal, url in enumerate(document_urls(value), start=1):
                document_key = (kind, url)
                if document_key in seen_documents:
                    continue
                seen_documents.add(document_key)
                documents.append({
                    "document_id": stable_id("doc", version_id, record.record_sha256, kind, url),
                    "material_number": record.material_number,
                    "document_type": kind,
                    "url": url,
                    "source_field": field_key,
                    "source_ordinal": ordinal,
                })
            continue

        normalized_header = normalize_alias(header.source_header)
        if any(marker in normalized_header for marker in SEMANTIC_EXCLUDED_HEADER_MARKERS):
            continue
        searchable_value = semantic_value(header.source_header, value)
        if searchable_value:
            grouped[field_group(field_key)].append((field_key, f"{header.source_header}: {searchable_value}"))

    chunks: list[dict[str, Any]] = []
    group_order = [group for group, _ in FIELD_GROUPS] + ["additional"]
    for group in group_order:
        lines = grouped.get(group, [])
        if lines:
            chunks.extend(pack_chunks(record, version_id, group, lines))

    return chunks, relationships, documents


def file_report(path: Path, root: Path) -> dict[str, object]:
    raw = path.read_bytes()
    return {
        "path": path.relative_to(root).as_posix(),
        "bytes": len(raw),
        "sha256": sha256_bytes(raw),
    }


def write_json(path: Path, value: object, *, compact: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    separators = (",", ":") if compact else None
    path.write_text(json.dumps(value, ensure_ascii=False, indent=None if compact else 2, separators=separators) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--manifest", required=True, type=Path)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    args = parser.parse_args()

    source = args.source.resolve()
    output_dir = args.output_dir.resolve()
    manifest_path = args.manifest.resolve()
    if not source.is_file():
        raise SystemExit(f"Workbook not found: {source}")

    output_dir.mkdir(parents=True, exist_ok=True)
    for pattern in (
        "master-catalog-stage-*.sql", "master-catalog-activate.sql", "master-catalog-qa.json",
        "master-catalog-field-dictionary.json", "master-catalog-records.ndjson",
        "master-catalog-chunks.ndjson", "master-catalog-vector-seed.ndjson",
    ):
        for old_path in output_dir.glob(pattern):
            old_path.unlink()

    source_bytes = source.read_bytes()
    source_sha256 = sha256_bytes(source_bytes)
    version_id = "mcv_" + sha256_text(
        f"{CATALOG_SCHEMA_VERSION}:{CATALOG_GENERATOR_VERSION}:{source_sha256}"
    )[:24]
    generated_at = utc_now()

    workbook = load_workbook(source, read_only=True, data_only=True)
    if args.sheet not in workbook.sheetnames:
        raise SystemExit(f"Workbook sheet not found: {args.sheet}")
    sheet = workbook[args.sheet]
    rows = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration as error:
        raise SystemExit("Workbook sheet is empty.") from error
    headers = disambiguate_headers(raw_headers)
    header_by_key = {header.field_key: header for header in headers}

    records: list[Record] = []
    missing_material_rows: list[int] = []
    material_rows: dict[str, int] = {}
    duplicate_materials: list[dict[str, object]] = []
    populated_counts: list[int] = []

    for source_row, values in enumerate(rows, start=2):
        fields = {
            header.field_key: json_value(values[header.ordinal - 1] if header.ordinal <= len(values) else None)
            for header in headers
            if header.ordinal <= len(values) and populated(values[header.ordinal - 1])
        }
        if not fields:
            continue
        material_number = record_field(fields, headers, "Material Number")
        if not material_number:
            missing_material_rows.append(source_row)
            continue
        if material_number in material_rows:
            duplicate_materials.append({
                "material_number": material_number,
                "first_source_row": material_rows[material_number],
                "duplicate_source_row": source_row,
            })
            continue
        material_rows[material_number] = source_row

        payload = {
            "schema_version": CATALOG_SCHEMA_VERSION,
            "material_number": material_number,
            "source": {"file": source.name, "sheet": args.sheet, "row": source_row},
            "fields": fields,
        }
        record_json = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        records.append(Record(
            material_number=material_number,
            product_name=record_field(fields, headers, "Material Description (Global English)"),
            parent_family=record_field(fields, headers, "Parent Family Name"),
            family=record_field(fields, headers, "Family Name"),
            trade_name=record_field(fields, headers, "Trade Name"),
            ai_summary=record_field(fields, headers, "AI_Summary"),
            ai_search_index=record_field(fields, headers, "AI_Search_Index"),
            source_row=source_row,
            fields=fields,
            record_json=record_json,
            record_sha256=sha256_text(record_json),
        ))
        populated_counts.append(len(fields))

    workbook.close()
    if duplicate_materials:
        raise SystemExit(f"Duplicate material numbers prevent staging: {len(duplicate_materials)} duplicate rows")
    if not records:
        raise SystemExit("No material records were found in the workbook.")

    material_numbers = {record.material_number for record in records}
    stage_writer = SqlShardWriter(output_dir, version_id)
    stage_writer.append(
        "INSERT OR IGNORE INTO master_catalog_versions("
        "version_id,schema_version,generator_version,source_file,source_sha256,source_bytes,source_sheet,source_rows,source_columns,status,generated_at"
        ") VALUES ("
        + ",".join(sql_literal(value) for value in (
            version_id, CATALOG_SCHEMA_VERSION, CATALOG_GENERATOR_VERSION,
            source.name, source_sha256, len(source_bytes),
            args.sheet, len(records), len(headers), "loading", generated_at,
        ))
        + ");"
    )

    counts: Counter[str] = Counter()
    category_counts: Counter[str] = Counter()
    family_counts: Counter[str] = Counter()
    named_parent_families: set[str] = set()
    named_families: set[str] = set()
    relationship_type_counts: Counter[str] = Counter()
    chunk_kind_counts: Counter[str] = Counter()
    blank_core = Counter()
    numeric_fields = Counter()
    chunk_lengths: list[int] = []
    chunk_samples: list[dict[str, object]] = []

    materials_insert = BatchedInsert(stage_writer, "master_materials", [
        "version_id", "material_number", "product_name", "parent_family", "family", "trade_name",
        "ai_summary", "ai_search_index", "source_row", "record_sha256", "record_json",
    ])
    segmented_records: list[Record] = []
    records_path = output_dir / "master-catalog-records.ndjson"
    with records_path.open("w", encoding="utf-8", newline="\n") as record_output:
        for record in records:
            material_row = (
                version_id, record.material_number, record.product_name, record.parent_family or None,
                record.family or None, record.trade_name or None, record.ai_summary or None,
                record.ai_search_index or None, record.source_row, record.record_sha256, record.record_json,
            )
            if not materials_insert.fits_single(material_row):
                base_json = json.dumps({
                    "schema_version": CATALOG_SCHEMA_VERSION,
                    "material_number": record.material_number,
                    "source": {"file": source.name, "sheet": args.sheet, "row": record.source_row},
                    "fields": {},
                }, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
                material_row = (*material_row[:-1], base_json)
                segmented_records.append(record)
            materials_insert.add(material_row)
            record_output.write(record.record_json + "\n")
            counts["materials"] += 1
            category_counts[record.parent_family or "Uncategorized"] += 1
            family_counts[record.family or "Uncategorized"] += 1
            if record.parent_family:
                named_parent_families.add(record.parent_family)
            if record.family:
                named_families.add(record.family)
            for label, value in (
                ("product_name", record.product_name), ("parent_family", record.parent_family),
                ("family", record.family), ("trade_name", record.trade_name),
                ("ai_summary", record.ai_summary), ("ai_search_index", record.ai_search_index),
            ):
                if not value:
                    blank_core[label] += 1
        materials_insert.flush()

    for record in segmented_records:
        for field_key, value in record.fields.items():
            json_path = f"$.fields.{field_key}"
            serialized_value = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
            statement = (
                "UPDATE master_materials SET record_json=json_set(record_json,"
                f"{sql_literal(json_path)},json({sql_literal(serialized_value)})) "
                f"WHERE version_id={sql_literal(version_id)} AND material_number={sql_literal(record.material_number)};"
            )
            if len(statement.encode("utf-8")) > MAX_INSERT_STATEMENT_BYTES:
                raise ValueError(
                    f"Field {field_key} for material {record.material_number} exceeds the conservative SQL statement limit"
                )
            stage_writer.append(statement)
        counts["segmented_record_json"] += 1

    aliases_insert = BatchedInsert(stage_writer, "master_aliases", [
        "alias_id", "version_id", "material_number", "alias_type", "alias", "normalized_alias",
    ])
    for record in records:
        seen_aliases: set[tuple[str, str]] = set()
        source_aliases: list[tuple[str, str]] = [
            ("material_number", record.material_number),
            ("trade_name", record.trade_name),
            ("product_name", record.product_name),
        ]
        for field_key, value in record.fields.items():
            if ALTERNATIVE_MODEL_HEADER.match(header_by_key[field_key].source_header):
                source_aliases.extend(
                    ("alternative_model", alias) for alias in split_alias_values(value)
                )
        for alias_type, alias in source_aliases:
            for normalized in identifier_alias_normalizations(alias_type, alias):
                key = (alias_type, normalized)
                if key in seen_aliases:
                    continue
                seen_aliases.add(key)
                aliases_insert.add((
                    stable_id("alias", version_id, record.material_number, alias_type, normalized),
                    version_id, record.material_number, alias_type, alias, normalized,
                ))
                counts["aliases"] += 1
    aliases_insert.flush()

    attributes_insert = BatchedInsert(stage_writer, "master_attributes", [
        "attribute_id", "version_id", "material_number", "field_key", "source_header",
        "source_column", "source_ordinal", "value_text", "value_number", "value_unit",
        "canonical_number", "canonical_unit", "value_json",
    ])
    for record in records:
        for field_key, value in record.fields.items():
            header = header_by_key[field_key]
            number, unit = parse_numeric(field_key, value)
            canonical_number, canonical_unit = canonical_measurement(number, unit, field_key)
            if number is not None:
                numeric_fields[field_key] += 1
                counts["numeric_attributes"] += 1
            if canonical_number is not None:
                counts["canonical_attributes"] += 1
            attributes_insert.add((
                stable_id("attr", version_id, record.record_sha256, field_key), version_id, record.material_number,
                field_key, header.source_header, header.column, header.ordinal, value_text(value),
                number, unit, canonical_number, canonical_unit,
                json.dumps(value, ensure_ascii=False, separators=(",", ":")),
            ))
            counts["attributes"] += 1
    attributes_insert.flush()

    relationships_insert = BatchedInsert(stage_writer, "master_relationships", [
        "relationship_id", "version_id", "source_material_number", "relationship_type",
        "target_material_number", "target_resolved", "source_field", "source_ordinal",
    ])
    documents_insert = BatchedInsert(stage_writer, "master_documents", [
        "document_id", "version_id", "material_number", "document_type", "url", "source_field", "source_ordinal",
    ])
    chunks_insert = BatchedInsert(stage_writer, "master_chunks", [
        "chunk_id", "version_id", "material_number", "parent_family", "family", "chunk_kind",
        "chunk_ordinal", "title", "content", "content_sha256", "field_keys_json", "metadata_json",
    ])
    chunks_path = output_dir / "master-catalog-chunks.ndjson"
    vector_seed_path = output_dir / "master-catalog-vector-seed.ndjson"
    with (
        chunks_path.open("w", encoding="utf-8", newline="\n") as chunk_output,
        vector_seed_path.open("w", encoding="utf-8", newline="\n") as vector_output,
    ):
        for record in records:
            chunks, relationships, documents = record_chunks(record, headers, material_numbers, version_id)
            for relationship in relationships:
                relationships_insert.add((
                    relationship["relationship_id"], version_id, relationship["source_material_number"],
                    relationship["relationship_type"], relationship["target_material_number"],
                    relationship["target_resolved"], relationship["source_field"], relationship["source_ordinal"],
                ))
                counts["relationships"] += 1
                counts["resolved_relationships" if relationship["target_resolved"] else "unresolved_relationships"] += 1
                relationship_type_counts[relationship["relationship_type"]] += 1
            for document in documents:
                documents_insert.add((
                    document["document_id"], version_id, document["material_number"], document["document_type"],
                    document["url"], document["source_field"], document["source_ordinal"],
                ))
                counts["documents"] += 1
            for chunk in chunks:
                metadata = {
                    "catalog_version": version_id,
                    "material_number": record.material_number,
                    "parent_family": record.parent_family or None,
                    "family": record.family or None,
                    "chunk_kind": chunk["chunk_kind"],
                    "source_file": source.name,
                    "source_sheet": args.sheet,
                    "source_row": record.source_row,
                }
                field_keys_json = json.dumps(chunk["field_keys"], ensure_ascii=False, separators=(",", ":"))
                metadata_json = json.dumps(metadata, ensure_ascii=False, separators=(",", ":"))
                chunks_insert.add((
                    chunk["chunk_id"], version_id, record.material_number, record.parent_family or None,
                    record.family or None, chunk["chunk_kind"], chunk["chunk_ordinal"], chunk["title"],
                    chunk["content"], chunk["content_sha256"], field_keys_json, metadata_json,
                ))
                chunk_record = {**chunk, "version_id": version_id, "metadata": metadata}
                chunk_output.write(json.dumps(chunk_record, ensure_ascii=False, separators=(",", ":")) + "\n")
                vector_output.write(json.dumps({
                    "id": chunk["chunk_id"],
                    "text": chunk["content"],
                    "namespace": f"master-{version_id}",
                    "metadata": metadata,
                }, ensure_ascii=False, separators=(",", ":")) + "\n")
                counts["chunks"] += 1
                chunk_kind_counts[chunk["chunk_kind"]] += 1
                chunk_lengths.append(len(chunk["content"]))
                if len(chunk_samples) < 8:
                    chunk_samples.append({
                        "chunk_id": chunk["chunk_id"],
                        "material_number": record.material_number,
                        "chunk_kind": chunk["chunk_kind"],
                        "characters": len(chunk["content"]),
                    })
    relationships_insert.flush()
    documents_insert.flush()
    chunks_insert.flush()

    stage_writer.append(
        "UPDATE master_catalog_versions SET "
        f"material_count={counts['materials']},alias_count={counts['aliases']},"
        f"attribute_count={counts['attributes']},relationship_count={counts['relationships']},"
        f"document_count={counts['documents']},chunk_count={counts['chunks']},"
        f"status=CASE WHEN status='loading' THEN 'staged' ELSE status END,staged_at={sql_literal(generated_at)} "
        f"WHERE version_id={sql_literal(version_id)};"
    )
    stage_writer.append(
        "UPDATE master_catalog_state SET "
        f"staged_version_id=CASE WHEN (SELECT status FROM master_catalog_versions WHERE version_id={sql_literal(version_id)})='staged' "
        f"THEN {sql_literal(version_id)} ELSE staged_version_id END,updated_at={sql_literal(generated_at)} "
        "WHERE singleton_id=1;"
    )
    stage_paths = stage_writer.close()

    activation_path = output_dir / "master-catalog-activate.sql"
    activation_path.write_text(
        "-- Catalog activation is intentionally API-gated. This compatibility\n"
        "-- artifact is non-mutating; use the authenticated catalog-admin activate\n"
        "-- action after count, vector, and retrieval-evaluation checks pass.\n"
        "SELECT " + sql_literal(
            f"Catalog {version_id} remains staged; activate it through the authenticated catalog-admin API."
        ) + " AS activation_notice;\n",
        encoding="utf-8",
    )

    duplicate_headers = [
        {
            "source_header": source_header,
            "occurrences": occurrence_count,
            "field_keys": [header.field_key for header in headers if header.source_header == source_header],
        }
        for source_header, occurrence_count in Counter(header.source_header for header in headers).items()
        if occurrence_count > 1
    ]
    field_dictionary_path = output_dir / "master-catalog-field-dictionary.json"
    write_json(field_dictionary_path, {
        "version_id": version_id,
        "source_file": source.name,
        "source_sheet": args.sheet,
        "fields": [header.__dict__ for header in headers],
    })

    errors = []
    if missing_material_rows:
        errors.append(f"{len(missing_material_rows)} populated rows have no material number and were not staged.")
    if any(length > MAX_CHUNK_CHARACTERS for length in chunk_lengths):
        errors.append("At least one generated chunk exceeds the configured character limit.")
    warnings = []
    if blank_core["parent_family"]:
        warnings.append(f"{blank_core['parent_family']} materials have no parent family.")
    if blank_core["family"]:
        warnings.append(f"{blank_core['family']} materials have no family.")
    if blank_core["trade_name"]:
        warnings.append(f"{blank_core['trade_name']} materials have no trade name; material number and product name remain searchable aliases.")
    if counts["unresolved_relationships"]:
        warnings.append(f"{counts['unresolved_relationships']} relationship targets do not resolve to a material in this workbook version.")

    qa_path = output_dir / "master-catalog-qa.json"
    qa = {
        "qa_version": "1.1.0",
        "status": "blocked" if errors else "review_required" if warnings else "ready",
        "generated_at": generated_at,
        "generator_version": CATALOG_GENERATOR_VERSION,
        "version_id": version_id,
        "source": {
            "file": source.name,
            "sheet": args.sheet,
            "sha256": source_sha256,
            "bytes": len(source_bytes),
            "rows_staged": len(records),
            "columns": len(headers),
        },
        "counts": dict(sorted(counts.items())),
        "category_counts": dict(sorted(category_counts.items())),
        "family_count": len(family_counts),
        "named_parent_family_count": len(named_parent_families),
        "named_family_count": len(named_families),
        "relationship_type_counts": dict(sorted(relationship_type_counts.items())),
        "chunk_kind_counts": dict(sorted(chunk_kind_counts.items())),
        "populated_fields_per_record": {
            "minimum": min(populated_counts),
            "maximum": max(populated_counts),
            "average": round(sum(populated_counts) / len(populated_counts), 2),
        },
        "numeric_field_count": len(numeric_fields),
        "numeric_attribute_count": counts["numeric_attributes"],
        "duplicate_headers": duplicate_headers,
        "missing_material_rows": missing_material_rows,
        "blank_core_fields": dict(sorted(blank_core.items())),
        "chunk_characters": {
            "limit": MAX_CHUNK_CHARACTERS,
            "minimum": min(chunk_lengths),
            "maximum": max(chunk_lengths),
            "average": round(sum(chunk_lengths) / len(chunk_lengths), 2),
            "violations": sum(length > MAX_CHUNK_CHARACTERS for length in chunk_lengths),
        },
        "errors": errors,
        "warnings": warnings,
    }
    write_json(qa_path, qa)

    output_reports = [file_report(path, output_dir) for path in [
        *stage_paths, activation_path, qa_path, field_dictionary_path, records_path, chunks_path, vector_seed_path,
    ]]
    manifest = {
        "manifest_version": "1.1.0",
        "catalog_schema_version": CATALOG_SCHEMA_VERSION,
        "generator": {
            "name": "scripts/import-master-catalog.py",
            "version": CATALOG_GENERATOR_VERSION,
            "alias_strategy": "source-plus-compact-alpha-digit-spaced",
        },
        "version_id": version_id,
        "status": "staged",
        "generated_at": generated_at,
        "source": {
            "file": source.name,
            "sheet": args.sheet,
            "sha256": source_sha256,
            "bytes": len(source_bytes),
            "rows": len(records),
            "columns": len(headers),
        },
        "counts": {
            "materials": counts["materials"],
            "aliases": counts["aliases"],
            "attributes": counts["attributes"],
            "numeric_attributes": counts["numeric_attributes"],
            "canonical_attributes": counts["canonical_attributes"],
            "relationships": counts["relationships"],
            "resolved_relationships": counts["resolved_relationships"],
            "unresolved_relationships": counts["unresolved_relationships"],
            "documents": counts["documents"],
            "chunks": counts["chunks"],
            "categories": len(category_counts),
            "families": len(family_counts),
            "named_parent_families": len(named_parent_families),
            "named_families": len(named_families),
        },
        "chunking": {
            "maximum_characters": MAX_CHUNK_CHARACTERS,
            "maximum_observed_characters": max(chunk_lengths),
            "vector_namespace": f"master-{version_id}",
            "deterministic_id_maximum": max(len(sample["chunk_id"]) for sample in chunk_samples),
        },
        "sql": {
            "shards": len(stage_paths),
            "conservative_statement_limit_bytes": MAX_INSERT_STATEMENT_BYTES,
            "maximum_observed_statement_bytes": stage_writer.maximum_statement_bytes,
        },
        "quality": {
            "status": qa["status"],
            "errors": len(errors),
            "warnings": len(warnings),
            "duplicate_header_groups": len(duplicate_headers),
            "missing_material_rows": len(missing_material_rows),
        },
        "migration": "migrations/0004_retrieval_evaluation_build_fingerprint.sql",
        "migrations": [
            "migrations/0001_master_catalog.sql",
            "migrations/0002_master_catalog_rollout.sql",
            "migrations/0003_reconcile_staged_master.sql",
            "migrations/0004_retrieval_evaluation_build_fingerprint.sql",
        ],
        "outputs": output_reports,
        "chunk_samples": chunk_samples,
    }
    write_json(manifest_path, manifest)
    print(json.dumps({
        "source": str(source),
        "version_id": version_id,
        "status": qa["status"],
        "materials": counts["materials"],
        "attributes": counts["attributes"],
        "relationships": counts["relationships"],
        "documents": counts["documents"],
        "chunks": counts["chunks"],
        "sql_shards": len(stage_paths),
        "output_dir": str(output_dir),
        "manifest": str(manifest_path),
    }, indent=2))


if __name__ == "__main__":
    main()
